from flask import Flask 
from flask import render_template
from flask import request
from flask import send_file
from flask import Flask, render_template, request, redirect, url_for,session
import cv2
import pandas as pd
import dlib
import sys
import sqlite3
import cognitive_face as CF
#from global_variables import personGroupId
import os, urllib
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import json
import shutil
import dotenv

import pymongo
from pymongo import MongoClient

dotenv.load_dotenv()
personGroupId = 'test1'    
UPLOAD_FOLDER="mlFiles"

app = Flask(__name__)
app.secret_key = os.getenv('sec_key')



cluster = MongoClient(os.getenv('mongoDet'))
db = cluster["database2"]
Testcollection = db["test"]
Usercollection = db["users"] 
Admincollection = db["admin_user"]
currentDate1 = time.strftime("%d-%m-%y")






def identify():

    currentDate = time.strftime("%d%m%y")
    wb = load_workbook("mlFiles/reports.xlsx")
    sheet = wb.get_sheet_by_name('Cse15')
    Key = os.getenv('API_key')
    CF.Key.set(Key)

    connect = connect = sqlite3.connect("mlFiles/Face-DataBase")
    c = connect.cursor()

    attend = [0 for i in range(60)]	

    
    directory = 'mlFiles/CroppedFaces'
    for filename in os.listdir(directory):
        if filename.endswith(".jpg"):
            imgurl = urllib.request.pathname2url(os.path.join(directory, filename))
            
            res = CF.face.detect(imgurl)
            if len(res) != 1:
                print("No face detected.")
                continue
            
            faceIds = []
            for face in res:
                faceIds.append(face['faceId'])
            res = CF.face.identify(faceIds, personGroupId)
            print(filename)
            print(res)
            for face  in res:
                if not face['candidates']:
                    print("Unknown")
                else:
                    personId = face['candidates'][0]['personId']
                    print(personId)
                    c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
                    row = c.fetchone()
                    attend[int(row[0])] += 1
                    print(int(row[0]))
                    print(row[1] + " recognized")
            time.sleep(6)
    count=0
    for row in range(2, sheet.max_row+1):
        rn = sheet.cell(row,1).value
        
        if rn is not None:
            rn=str(rn)
            print("################")
            print(rn)
        if rn is not None:
            rn = rn[-2:]

            j=int(sheet.cell(row=row,column=6).value)
            j+=1
            sheet.cell(row=row,column=6).value=j 
            
            sheet.cell(row=row,column=4).value = currentDate
            if attend[int(rn)] != 0:
                count+=1
                sheet.cell(row=row,column=3).value = 1
                
                k=int(sheet.cell(row=row,column=5).value)
                k+=1
                sheet.cell(row=row,column=5).value=k 
    print(count)
    session["todayAttendance"]=count
    wb.save("mlFiles/reports.xlsx")	
    # Read and store content
    # of an excel file 
    read_file = pd.read_excel ("mlFiles/reports.xlsx")
  
    # Write the dataframe object
    # into csv file
    read_file.to_csv ("mlFiles/reports_csv.csv", 
                  index = None,
                  header=True)
    
    df=pd.read_csv("mlFiles/reports_csv.csv")
    data=df.to_dict("records")
    
    Testcollection.insert_many(data)
    print("#################")
    print("# data exported #")
    print("#################")
    print(session["user"])
    result= Admincollection.update_one({"username":session["user"]},{"$inc":{"workingDays":1}})
    print(result)
    session["workingday"]=Admincollection.find_one({"username":session["user"]})["workingDays"]
def detect(image_path):
    detector = dlib.get_frontal_face_detector()
    img = cv2.imread(image_path)
    dets = detector(img, 1)
    if  os.path.exists('mlFiles/CroppedFaces'):
        shutil.rmtree('mlFiles/CroppedFaces')
    if not os.path.exists('mlFiles/CroppedFaces'):
        
        os.makedirs('mlFiles/CroppedFaces')
    print("detected = " + str(len(dets)))
    for i, d in enumerate(dets):
            cv2.imwrite('mlFiles/CroppedFaces/face' + str(i + 1) + '.jpg', img[d.top():d.bottom(), d.left():d.right()])
    

@app.route("/predict",methods=["GET","POST"])
def predict():
    print(session["user"])
    if(session["user"]):
        if request.method == "POST":
            image_file = request.files["fileup"]
            
            if image_file:
                image_location=os.path.join(UPLOAD_FOLDER,image_file.filename)
                image_file.save(image_location)
                detect(image_location)
                identify()
                return redirect(url_for('adHome'))
    else:
        return redirect(url_for('SignIn'))
            
    return render_template("upload.html")

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/index')
def index2():
    return render_template('index2.html')

@app.route('/Logout')
def logout():
    clear()
    return redirect(url_for("signIn"))

@app.route('/phase2')
def phase2():
    return render_template('phase2.html')

@app.route('/contactUs')
def contactUs():
    return render_template('contactUs.html')




     
@app.route('/')
def main():
    clear()
    return render_template('index.html')
@app.route('/home')
def home():
    return render_template('index.html')


@app.route('/team')
def team():
    return render_template('team.html')

@app.route('/studHome',methods=['POST','GET'])
def studHome():
    if request.method == 'POST'or'GET':
        if "user" in session:
            user=session["user"]
            thour=session["totalhour"]
            tattendance=session["totalattendance"]
            percent=int(thour/tattendance*100)
            return render_template('student.html',Date=currentDate1, User=user,Thour=thour ,Tattend=tattendance,Percent=percent)
        
        return redirect(url_for('signIn')) 
    else:
        return redirect(url_for('signIn'))
    



@app.route('/adHome',methods=['POST','GET'])
def adHome():
    if request.method == 'POST'or'GET':
        if "user" in session:
            user=session["user"]
            studno=session["studno"]
            workingdays=session["workingday"]
            if "todayAttendance" in session:
                attendances=session["todayAttendance"]
            else:
                attendances="Not Taken"
        
            return render_template('adminhome.html',Date=currentDate1, User=user,StudNo=studno,Wday=workingdays, Tattend=attendances)
        
        return redirect(url_for('home')) 
    else:
         return render_template('signin.html')
    

     
@app.route('/register',methods=['POST','GET'])

def signUp():     
    
    if request.method == 'POST':

        _name = request.form['username3']
        _password = request.form['password3']
        # validate the received values
        users = Usercollection.find_one({"username":_name})
        if (users):
            return render_template("register.html",messagebox="../static/js/alertExist.js")
                
        else:
            usersave = {"username":_name, "password":_password, "id":"student"}
            Usercollection.insert_one(usersave)
            return render_template("register.html",messagebox="../static/js/alertAdded.js")
                
    else:
        return render_template("register.html")


 
@app.route('/signIn', methods=['GET', 'POST'])
def signIn():
    
    if request.method == 'POST':
        # Get Form Fields
        _username = request.form['username1'] 
        _password = request.form['password1']      
        # Get user by usernane
        users = Usercollection.find_one({"username":_username}) 
        session["user"]=users["username"]
        
        print(users) # result 1
        
        if(users):
            password_user=(users["password"])
            if(password_user==_password):
                if (users["id"]=="admin"):
                    users2 = Admincollection.find_one({"username":_username})
                    print(users2["username"])
                    
                    session["studno"]=users2["studNo"]
                    session["workingday"]=users2["workingDays"]
                    return redirect(url_for('adHome'))  

                else:
                    k=0
                    users2 = Admincollection.find({"id":"admin"})
                    for result in users2:
                        k+=result["workingDays"]
                    session["totalhour"]=k
                    
                    users = Testcollection.find({"Name":"Bevan V Thomas"})

                    for result in users:
                        k=(result["TotalAttendace"])
                    session["totalattendance"]=k

                    return redirect(url_for('studHome'))   
           
            else:
                clear()
                return render_template('signin.html',messagebox="../static/js/alertPassword.js")
            
        else:
            return render_template('signin.html',messagebox="../static/js/alertUser.js")
                
    else:
        return render_template('signin.html')
            

def clear():
    print("sessions cleared")
    if "user" in session:  
        print("user : ",session["user"])
        session.pop("user", None)     
        

    if "studno" in session: 
        print("studno : ",session["studno"])
        session.pop("studno", None)

    if "workingday" in session: 
        print("workingday : ",session["workingday"])
        session.pop("workingday", None)

    if "todayAttendance" in session: 
        print("todayAttendance : ",session["todayAttendance"])
        session.pop("todayAttendance", None)

    
    





if __name__ == '__main__':
    app.run(debug=True)
